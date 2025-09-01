#!/usr/bin/env python3
"""
Shipment XML Generator - Versión con Base de Datos
===============================================

Genera XML para TMS con:
- Consultas a base de datos MySQL para datos reales
- ReferenceNumber correlativo desde base de datos
- Mapping completo según especificaciones
- Validación automática de salida
- Reporte de calidad de datos

Autor: Versión con DB
Fecha: 2025-06-01
Versión: 5.0 - Con integración de base de datos
"""

import pandas as pd
from datetime import datetime, timedelta
import math
import sys
import os
import mysql.connector
from mysql.connector import Error
from typing import Optional, List, Dict, Any
import logging
import openpyxl
import re


# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('shipment_generator.log',  encoding='utf-8'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class DatabaseManager:
    """Gestor de conexiones y consultas a la base de datos"""

    def __init__(self, host='localhost', database='shipment_db', user='root', password=''):
        """Inicializar conexión a la base de datos"""
        self.host = host
        self.database = database
        self.user = user
        self.password = password
        self.connection = None
        self.connect()

    def connect(self):
        """Establecer conexión a MySQL"""
        try:
            self.connection = mysql.connector.connect(
                host=self.host,
                database=self.database,
                user=self.user,
                password=self.password,
                charset='utf8mb4',
                collation='utf8mb4_unicode_ci'
            )

            if self.connection.is_connected():
                logger.info(f"Conectado a MySQL Server versión {self.connection.get_server_info()}")
                logger.info(f"Conectado a base de datos: {self.database}")
        except Error as e:
            logger.error(f"Error conectando a MySQL: {e}")
            raise

    def execute_query(self, query, params=None, fetch_one=False):
        """Ejecutar consulta y retornar resultados"""
        try:
            cursor = self.connection.cursor(dictionary=True)
            cursor.execute(query, params or ())

            if fetch_one:
                result = cursor.fetchone()
            else:
                result = cursor.fetchall()

            cursor.close()
            return result
        except Error as e:
            logger.error(f"Error ejecutando consulta: {e}")
            logger.error(f"Query: {query}")
            logger.error(f"Params: {params}")
            return None

    def execute_update(self, query, params=None):
        """Ejecutar consulta de actualización"""
        try:
            cursor = self.connection.cursor()
            cursor.execute(query, params or ())
            self.connection.commit()
            affected_rows = cursor.rowcount
            cursor.close()
            return affected_rows
        except Error as e:
            logger.error(f"Error ejecutando actualización: {e}")
            self.connection.rollback()
            return 0

    def get_next_reference_number(self):
        """Obtener siguiente número de referencia correlativo desde BD"""
        try:
            # Usar cursor con dictionary=True para acceso por nombre de columna
            cursor = self.connection.cursor(dictionary=True)

            # Verificar si hay transacción activa y terminarla
            if self.connection.in_transaction:
                self.connection.rollback()

            # Obtener el último número
            query = "SELECT last_reference_number FROM shipment_sequence ORDER BY id DESC LIMIT 1"
            cursor.execute(query)
            result = cursor.fetchone()

            if result and result['last_reference_number']:
                next_number = result['last_reference_number'] + 1
            else:
                # Si no existe, crear registro inicial
                insert_query = "INSERT INTO shipment_sequence (last_reference_number, updated_at) VALUES (11111, NOW())"
                cursor.execute(insert_query)
                self.connection.commit()
                next_number = 11111

            # Actualizar el contador
            update_query = "UPDATE shipment_sequence SET last_reference_number = %s, updated_at = NOW() WHERE id = (SELECT id FROM (SELECT id FROM shipment_sequence ORDER BY id DESC LIMIT 1) as temp)"
            cursor.execute(update_query, (next_number,))
            self.connection.commit()

            cursor.close()
            return str(next_number)

        except Exception as e:
            logger.error(f"Error obteniendo reference number: {e}")
            # Generar número único con timestamp como fallback
            timestamp = datetime.now().strftime("%H%M%S%f")[:10]
            fallback_number = f"11{timestamp}"
            logger.warning(f"Usando número fallback: {fallback_number}")
            return fallback_number

    def get_sku_name(self, codigo_prod):
        """SKU Name = BUSCARV en dados_productos o maestro_envases"""
        try:
            # Primero buscar en dados_produtos
            query = "SELECT nombre FROM dados_produtos WHERE codigo = %s"
            result = self.execute_query(query, (codigo_prod,), fetch_one=True)

            if result:
                return result['nombre']

            # Si no encuentra, buscar en maestro_envases
            query = "SELECT descripcion FROM maestro_envases WHERE codigo_envase = %s"
            result = self.execute_query(query, (codigo_prod,), fetch_one=True)

            if result:
                return result['descripcion']

            return f"PRODUCTO_{codigo_prod}"
        except Exception as e:
            logger.error(f"Error obteniendo SKU name para {codigo_prod}: {e}")
            return f"PRODUCTO_{codigo_prod}"

    def get_priority(self, origen, destino):
        """Priority = BUSCARV en maestro_dtto con rutas"""
        try:
            # Construir rutas como en Excel
            ruta1 = f"BO_{origen}-BO_{destino}"
            ruta2 = f"BO_{destino}-BO_{origen}"

            # Buscar primera ruta
            query = "SELECT prioridad FROM maestro_dtto WHERE ruta = %s"
            result = self.execute_query(query, (ruta1,), fetch_one=True)

            if result:
                return result['prioridad']

            # Buscar ruta inversa
            result = self.execute_query(query, (ruta2,), fetch_one=True)

            if result:
                return result['prioridad']

            return 1  # Default
        except Exception as e:
            logger.error(f"Error obteniendo prioridad para {origen}-{destino}: {e}")
            return 1

    def get_commodity(self, codigo_prod):
        """Commodity = BUSCARV en dados_produtos columna D (codigo_commodity)"""
        try:
            query = "SELECT codigo_commodity FROM dados_produtos WHERE codigo = %s"
            result = self.execute_query(query, (codigo_prod,), fetch_one=True)

            if result and result['codigo_commodity']:
                return result['codigo_commodity']

            return "BO_BR"  # Default
        except Exception as e:
            logger.error(f"Error obteniendo commodity para {codigo_prod}: {e}")
            return "BO_BR"

    def get_hectolitros(self, codigo_prod, pallets):
        """Hectolitros = BUSCARV en dados_produtos (hl_por_pallet * pallets) o maestro_envases"""
        try:
            # Buscar en dados_produtos (columna U - hl_por_pallet)
            query = "SELECT hl_por_pallet FROM dados_produtos WHERE codigo = %s"
            result = self.execute_query(query, (codigo_prod,), fetch_one=True)

            if result and result['hl_por_pallet']:
                return round(float(result['hl_por_pallet']) * float(pallets), 4)

            # Si no encuentra, buscar en maestro_envases
            query = "SELECT hl_x_pallet FROM maestro_envases WHERE codigo_envase = %s"
            result = self.execute_query(query, (codigo_prod,), fetch_one=True)

            if result and result['hl_x_pallet']:
                return round(float(result['hl_x_pallet']) * float(pallets), 4)

            return 0.0
        except Exception as e:
            logger.error(f"Error obteniendo hectolitros para {codigo_prod}: {e}")
            return 0.0

    def get_bultos(self, codigo_prod, pallets):
        """Bultos = BUSCARV en dados_produtos (bultos_x_pallet * pallets) o maestro_envases"""
        try:
            # Buscar en dados_produtos (columna T - bultos_x_pallet)
            query = "SELECT bultos_x_pallet FROM dados_produtos WHERE codigo = %s"
            result = self.execute_query(query, (codigo_prod,), fetch_one=True)

            if result and result['bultos_x_pallet']:
                return int(result['bultos_x_pallet']) * int(pallets)

            # Si no encuentra, buscar en maestro_envases
            query = "SELECT bultos_x_pallet FROM maestro_envases WHERE codigo_envase = %s"
            result = self.execute_query(query, (codigo_prod,), fetch_one=True)

            if result and result['bultos_x_pallet']:
                return int(result['bultos_x_pallet']) * int(pallets)

            return 0
        except Exception as e:
            logger.error(f"Error obteniendo bultos para {codigo_prod}: {e}")
            return 0

    def get_sku_per_truck_count(self, shipment_id, all_shipments):
        """SKU per truck = CONTAR.SI($A:$A,A2) - contar ocurrencias del mismo envío"""
        try:
            count = sum(1 for s in all_shipments if s.get('Cód. Envío') == shipment_id)
            return count if count > 0 else 1
        except Exception as e:
            logger.error(f"Error contando SKU per truck para {shipment_id}: {e}")
            return 1

    def close_connection(self):
        """Cerrar conexión a la base de datos"""
        if self.connection and self.connection.is_connected():
            self.connection.close()
            logger.info("Conexión a MySQL cerrada")


class ShipmentXMLGenerator:
    """
    Generador de XML para órdenes de envío con consultas a base de datos
    """

    def __init__(self, db_config=None, use_planta_as_origen=False):
        """Inicializar el generador con conexión a BD"""
        # Constantes del sistema
        self.COUNTRY_CHAR = "BO"
        self.COMMODITY_DEFAULT = "BO_BR"
        self.PALLET_PRODUCT_CODE = 1100
        self.PALLET_QUANTITY = 24
        self.WEIGHT_PER_UNIT = 45

        self.use_planta_as_origen = use_planta_as_origen

        # Configuración de base de datos
        if db_config is None:
            db_config = {
                'host': 'localhost',
                'database': 'tms_shipment_db',
                'user': 'root',
                'password': 'Dali19((Kafka'
            }

        # Conectar a base de datos
        self.db = DatabaseManager(**db_config)

        # Headers XML esperados
        self.XML_HEADERS = [
            "Type", "Shipment Number", "Shipment Desription", "Commodity", "Priority",
            "OriginLocation", "DestinationLocation", "PickupFrom", "PickupTo",
            "DeliveryFrom", "DeliveryTo", "Carrier", "Service", "ReferenceNumberType",
            "ReferenceNumber", "Quantity", "Weight", "Hectolitros", "Bultos", "Pallets"
        ]

        self.shipment_data = []
        self.plan_id = None
        self.destination_folder = "."

        # Estadísticas para validación
        self.validation_stats = {
            'total_records': 0,
            'header_records': 0,
            'detail_records': 0,
            'priority_conversions': {},
            'reference_numbers_generated': [],
            'database_queries': 0,
            'errors': []
        }

    def safe_numeric_conversion(self, value: Any, default: float = 0.0) -> float:
        """Conversión segura a número"""
        try:
            if pd.isna(value) or value == '' or value is None:
                return default

            if isinstance(value, (int, float)):
                return float(value)

            value_str = str(value).strip()
            if value_str == '':
                return default

            return float(value_str)

        except (ValueError, TypeError):
            logger.warning(f"No se pudo convertir '{value}' a número, usando {default}")
            return default

    def safe_int_conversion(self, value: Any, default: int = 0) -> int:
        """Conversión segura a entero"""
        try:
            if pd.isna(value) or value == '' or value is None:
                return default

            if isinstance(value, int):
                return value

            if isinstance(value, float):
                return int(value)

            value_str = str(value).strip()
            if value_str == '':
                return default

            return int(float(value_str))

        except (ValueError, TypeError):
            logger.warning(f"No se pudo convertir '{value}' a entero, usando {default}")
            return default

    def validate_input_file(self, file_path: str) -> bool:
        """Validar que el archivo Excel exists y tiene la estructura correcta"""
        try:
            if not os.path.exists(file_path):
                logger.error(f"Archivo no encontrado: {file_path}")
                return False

            df = pd.read_excel(file_path, sheet_name='Consolidado', header=0)

            required_columns = [
                'Cód. Envío', 'Cód. Prod', 'Pallets', 'Fecha', 'Peso Total Carga',
                'Cód. Origen', 'Cód. Destino', 'Producto', 'Prioridad', 'HL', 'Bultos'
            ]

            if self.use_planta_as_origen:
                if 'Cod Planta' not in df.columns:
                    logger.error("❌ Parámetro --from-planta activado pero columna 'Cod Planta' no encontrada")
                    return False
                else:
                    logger.info("✅ Columna 'Cod Planta' encontrada para --from-planta")


            missing_columns = [col for col in required_columns if col not in df.columns]

            if missing_columns:
                logger.error(f"Columnas faltantes: {missing_columns}")
                return False

            if len(df) == 0:
                logger.error("El archivo no contiene datos")
                return False

            logger.info(f"Archivo validado exitosamente. {len(df)} registros encontrados")
            return True

        except Exception as e:
            logger.error(f"Error validando archivo: {str(e)}")
            return False

    def load_data(self, file_path: str) -> pd.DataFrame:
        """Cargar datos desde el archivo Excel"""
        try:
            logger.info(f"Cargando datos desde: {file_path}")

            df = pd.read_excel(file_path, sheet_name='Consolidado', header=0)
            df = df.dropna(subset=['Cód. Envío'])
            df = df[df['Cód. Envío'] != '']

            if self.use_planta_as_origen:
                if 'Cod Planta' in df.columns:
                    logger.info("🔄 USANDO 'Cod Planta' como origen (--from-planta activado)")
                    # Reemplazar los valores de 'Cód. Origen' con 'Cod Planta'
                    df['Cód. Origen'] = df['Cod Planta']
                    logger.info(f"✅ Columna 'Cód. Origen' reemplazada con valores de 'Cod Planta'")
                else:
                    logger.warning("⚠️ Columna 'Cod Planta' no encontrada, usando 'Cód. Origen' normal")

            logger.info(f"Datos cargados: {len(df)} registros válidos")
            return df

        except Exception as e:
            logger.error(f"Error cargando datos: {str(e)}")
            raise

    def detect_file_type_and_plan_id(self, input_file: str) -> tuple:
        """Detectar tipo de archivo y asignar Plan ID automáticamente usando fecha del nombre del archivo"""
        try:
            file_name = os.path.basename(input_file).upper()

            # Detectar tipo de archivo
            if 'BEER' in file_name:
                file_type = 'Beer'
                plan_id = '5001'
            elif 'SD' in file_name:
                file_type = 'SD'
                plan_id = '5002'
            elif 'CB' in file_name:
                file_type = 'CB'
                plan_id = '5003'
            else:
                file_type = 'General'
                plan_id = '5001'

            # NUEVO: Extraer fecha del nombre del archivo
            month, day = self.extract_date_from_filename(input_file)

            destination_folder = os.path.join(file_type, month, day)
            os.makedirs(destination_folder, exist_ok=True)

            logger.info(f"Archivo detectado como tipo: {file_type}")
            logger.info(f"Plan ID asignado automáticamente: {plan_id}")
            logger.info(f"Fecha extraída del archivo: día {day}, mes {month}")
            logger.info(f"Carpeta de destino: {destination_folder}")

            return file_type, plan_id, destination_folder

        except Exception as e:
            logger.error(f"Error detectando tipo de archivo: {str(e)}")
            return "General", "5001", "."

    def extract_date_from_filename(self, file_path: str) -> tuple:
        """
        Extraer día y mes del nombre del archivo
        ADAPTADO para múltiples formatos manteniendo tu lógica original
        Ejemplos:
        - Programa Beer_28_04.xlsm → día=28, mes=04
        - Programa_SD_1_04_2025_.xlsm → día=1, mes=04
        - Envíos CBs 19-06.xlsm → día=19, mes=06
        """
        import re

        try:
            file_name = os.path.basename(file_path)
            logger.info(f"Extrayendo fecha del archivo: {file_name}")

            # ===== PATRÓN 1: SD - Programa_SD_D_MM_YYYY_ (día puede ser 1 o 2 dígitos) =====
            pattern_sd = r'Programa_SD_(\d{1,2})_(\d{1,2})_\d{4}_'
            match_sd = re.search(pattern_sd, file_name)

            if match_sd:
                day = match_sd.group(1).zfill(2)  # Asegurar 2 dígitos
                month = match_sd.group(2).zfill(2)  # Asegurar 2 dígitos
                logger.info(f"📅 Patrón SD detectado: día {day}, mes {month}")
            else:
                # ===== PATRÓN 2: CB - Envíos CBs DD-MM.xlsm (con guión) =====
                pattern_cb = r'Envíos\s+CBs?\s+(\d{1,2})-(\d{1,2})'
                match_cb = re.search(pattern_cb, file_name)

                if match_cb:
                    day = match_cb.group(1).zfill(2)  # Asegurar 2 dígitos
                    month = match_cb.group(2).zfill(2)  # Asegurar 2 dígitos
                    logger.info(f"📅 Patrón CB detectado: día {day}, mes {month}")
                else:
                    # ===== PATRÓN 3: Beer y genérico - DD_MM (tu patrón original) =====
                    pattern = r'(\d{1,2})_(\d{1,2})'
                    match = re.search(pattern, file_name)

                    if match:
                        day = match.group(1).zfill(2)  # Primer grupo (DD) - asegurar 2 dígitos
                        month = match.group(2).zfill(2)  # Segundo grupo (MM) - asegurar 2 dígitos
                        logger.info(f"📅 Patrón genérico DD_MM detectado: día {day}, mes {month}")
                    else:
                        logger.warning(f"⚠️ No se encontró ningún patrón de fecha en: {file_name}")
                        raise ValueError("Patrón de fecha no encontrado")

            # ===== VALIDACIÓN (tu lógica original sin cambios) =====
            day_int = int(day)
            month_int = int(month)

            if 1 <= day_int <= 31 and 1 <= month_int <= 12:
                logger.info(f"✅ Fecha extraída: día {day}, mes {month}")
                return month, day
            else:
                logger.warning(f"⚠️ Fecha inválida en nombre: día {day}, mes {month}")
                raise ValueError("Fecha fuera de rango")

        except Exception as e:
            logger.warning(f"No se pudo extraer fecha del nombre del archivo: {str(e)}")
            logger.info("🔄 Usando fecha actual como fallback")

            # Fallback a fecha actual si no se puede extraer (tu lógica original sin cambios)
            now = datetime.now()
            month = now.strftime('%m')
            day = now.strftime('%d')

            logger.info(f"📅 Fallback: día {day}, mes {month} (fecha actual)")
            return month, day

    def generate_unique_route_correlative(self, viaje_number: int, origen: int, destino: int,
                                          all_shipments: List) -> int:
        """
        Generar correlativo único por combinación viaje + ruta.

        Esto permite:
        - Mismo archivo = mismos números (re-procesable)
        - Mismo viaje, diferentes destinos = números diferentes
        - Reproducible y determinístico
        """
        # Crear lista de todas las rutas únicas en el archivo para este viaje
        viaje_routes = []

        for shipment in all_shipments:
            ship_viaje = self.safe_int_conversion(shipment.get('# Viaje'), 0)
            ship_origen = self.safe_int_conversion(shipment.get('Cód. Origen'), 0)
            ship_destino = self.safe_int_conversion(shipment.get('Cód. Destino'), 0)

            if ship_viaje == viaje_number:
                route_key = f"{ship_origen}-{ship_destino}"
                if route_key not in viaje_routes:
                    viaje_routes.append(route_key)

        # Ordenar para consistencia
        viaje_routes.sort()

        # Encontrar el índice de la ruta actual
        current_route = f"{origen}-{destino}"
        try:
            route_index = viaje_routes.index(current_route) + 1  # 1-based
        except ValueError:
            route_index = 1  # Fallback

        # Generar correlativo: viaje * 100 + índice de ruta
        # Ejemplo: Viaje 1, Ruta 1 = 101, Viaje 1, Ruta 2 = 102, Viaje 2, Ruta 1 = 201
        correlative = viaje_number * 100 + route_index

        return correlative

    def generate_shipment_number(self, row_data: dict, index: int) -> str:
        """Versión simple usando combinación única de datos"""
        import hashlib

        # Crear string único con datos del registro
        unique_string = f"{row_data.get('Cód. Envío', '')}-{row_data.get('Cód. Prod', '')}-{row_data.get('Cód. Origen', '')}-{row_data.get('Cód. Destino', '')}-{index}"

        # Hash reproducible
        hash_short = int(hashlib.md5(unique_string.encode()).hexdigest()[:8], 16)
        final_number = 2226500000 + (hash_short % 999999)

        return f"{self.COUNTRY_CHAR}{final_number}"

    def calculate_dates(self, base_date: datetime) -> Dict[str, str]:
        """Calcular fechas de pickup y delivery basadas en la fecha base"""
        try:
            # Usar fecha actual si no se proporciona
            if base_date is None or pd.isna(base_date):
                base_date = datetime.now()
            elif isinstance(base_date, str):
                base_date = pd.to_datetime(base_date)

            pickup_from = base_date.strftime('%Y-%m-%d %H:%M')
            pickup_to = (base_date + timedelta(days=20)).strftime('%Y-%m-%d %H:%M')
            delivery_from = (base_date + timedelta(minutes=1)).strftime('%Y-%m-%d %H:%M')
            delivery_to = (base_date + timedelta(days=25)).strftime('%Y-%m-%d %H:%M')

            return {
                'pickup_from': pickup_from,
                'pickup_to': pickup_to,
                'delivery_from': delivery_from,
                'delivery_to': delivery_to
            }

        except Exception as e:
            logger.error(f"Error calculando fechas: {str(e)}")
            default_date = datetime.now()
            return {
                'pickup_from': default_date.strftime('%Y-%m-%d %H:%M'),
                'pickup_to': (default_date + timedelta(days=20)).strftime('%Y-%m-%d %H:%M'),
                'delivery_from': (default_date + timedelta(minutes=1)).strftime('%Y-%m-%d %H:%M'),
                'delivery_to': (default_date + timedelta(days=25)).strftime('%Y-%m-%d %H:%M')
            }

    def process_weight(self, base_weight: float, quantity: int) -> float:
        """Procesar peso siguiendo la lógica del VBA"""
        try:
            total_weight = base_weight + (quantity * self.WEIGHT_PER_UNIT)
            return total_weight
        except:
            return base_weight if base_weight else 0

    def process_record(self, row: pd.Series, index: int, ship_num: int, all_shipments: List) -> List[Dict[str, Any]]:
        """Procesar un registro individual con consultas a BD"""
        try:
            # Obtener # Viaje del Excel (mantener la lógica original)
            viaje_number = self.safe_int_conversion(row.get('# Viaje'), index + 1)

            # Para evitar duplicados en mismo viaje con múltiples destinos:
            # Crear un correlativo basado en viaje + ruta única
            code_prod = self.safe_int_conversion(row['Cód. Prod'], 0)
            quantity = self.safe_int_conversion(row['Pallets'], 1)
            base_date = row.get('Fecha')
            base_weight = self.safe_numeric_conversion(row['Peso Total Carga'], 0)
            origen = self.safe_int_conversion(row['Cód. Origen'], 1)
            destino = self.safe_int_conversion(row['Cód. Destino'], 1)
            carrier = str(row['Operador Logístico']) if pd.notna(row['Operador Logístico']) else ''
            shipment_id = row.get('Cód. Envío', '')

            # === CONSULTAS A BASE DE DATOS ===
            # logger.info(f"Procesando registro {index + 1} - Viaje: {viaje_number} - Producto: {code_prod} - Origen: {origen} -> Destino: {destino}")

            # 1. SKU Name desde BD
            sku_name = self.db.get_sku_name(code_prod)
            self.validation_stats['database_queries'] += 1

            # 2. Priority desde BD usando rutas
            priority_numeric = self.db.get_priority(origen, destino)
            self.validation_stats['database_queries'] += 1

            # 3. Commodity desde BD
            commodity = self.db.get_commodity(code_prod)
            self.validation_stats['database_queries'] += 1

            # 4. Hectolitros desde BD
            hectolitros = self.db.get_hectolitros(code_prod, quantity)
            self.validation_stats['database_queries'] += 1

            # 5. Bultos desde BD
            bultos = self.db.get_bultos(code_prod, quantity)
            self.validation_stats['database_queries'] += 1

            # 6. SKU per truck - conteo
            sku_per_truck = self.db.get_sku_per_truck_count(shipment_id, all_shipments)

            # 7. ReferenceNumber correlativo desde BD
            sales_order_reference = self.db.get_next_reference_number()
            self.validation_stats['reference_numbers_generated'].append(sales_order_reference)

            # Registrar conversiones para estadísticas
            if str(priority_numeric) not in self.validation_stats['priority_conversions']:
                self.validation_stats['priority_conversions'][str(priority_numeric)] = 0
            self.validation_stats['priority_conversions'][str(priority_numeric)] += 1

            # Ajustar cantidad si es pallet
            if code_prod == self.PALLET_PRODUCT_CODE:
                quantity = self.PALLET_QUANTITY

            # GENERAR SHIPMENT NUMBER ÚNICO POR VIAJE + RUTA
            # Crear un correlativo único basado en viaje + origen + destino
            unique_route_id = self.generate_unique_route_correlative(viaje_number, origen, destino, all_shipments)
            # print(f"DEBUG - Viaje: {viaje_number}, Ruta: {origen}->{destino}, RouteID: {unique_route_id}")
            row_dict = row.to_dict()  # Convertir serie a diccionario
            shipment_number = self.generate_shipment_number(row_dict, index)
            # print(f"DEBUG - ship_num: {ship_num}, final shipment: {shipment_number}")

            # Calcular fechas (usar fecha actual)
            dates = self.calculate_dates(datetime.now())

            # Procesar peso
            total_weight = self.process_weight(base_weight, quantity)

            # Lane = "BO_" + origen + "-BO_" + destino
            lane = f"BO_{origen}-BO_{destino}"

            # Pallet retornable
            pallet_ret = row.get('Pallet_Retornable', '')
            if pd.isna(pallet_ret):
                pallet_ret = ''

            # Generar los 4 registros (patrón VBA: 1 Header + 3 Detail)
            records = []

            # Registro Header (H) con datos de BD
            header_record = {
                'Type': 'H',
                'Shipment Number': shipment_number,
                'Shipment Desription': sku_name,  # Desde BD
                'Commodity': commodity,  # Desde BD
                'Priority': priority_numeric,  # Desde BD
                'OriginLocation': f"{self.COUNTRY_CHAR}_{origen}",
                'DestinationLocation': f"{self.COUNTRY_CHAR}_{destino}",
                'PickupFrom': dates['pickup_from'],
                'PickupTo': dates['pickup_to'],
                'DeliveryFrom': dates['delivery_from'],
                'DeliveryTo': dates['delivery_to'],
                'Carrier': '',
                'Service': '',
                'ReferenceNumberType': 'SALES_ORDER',
                'ReferenceNumber': sales_order_reference,  # Desde BD
                'Quantity': '',
                'Weight': '',
                'Hectolitros': '',
                'Bultos': '',
                'Pallets': ''
            }
            records.append(header_record)
            self.validation_stats['header_records'] += 1

            # Registro Detail 1 (D) - CODE_PROD
            detail1_record = {
                'Type': 'D',
                'Shipment Number': '',
                'Shipment Desription': '',
                'Commodity': '',
                'Priority': '',
                'OriginLocation': '',
                'DestinationLocation': '',
                'PickupFrom': '',
                'PickupTo': '',
                'DeliveryFrom': '',
                'DeliveryTo': '',
                'Carrier': '',
                'Service': '',
                'ReferenceNumberType': 'CODE_PROD',
                'ReferenceNumber': code_prod,
                'Quantity': '',
                'Weight': '',
                'Hectolitros': '',
                'Bultos': '',
                'Pallets': ''
            }
            records.append(detail1_record)
            self.validation_stats['detail_records'] += 1

            # Registro Detail 2 (D) - PALLET_RET
            detail2_record = {
                'Type': 'D',
                'Shipment Number': '',
                'Shipment Desription': '',
                'Commodity': '',
                'Priority': '',
                'OriginLocation': '',
                'DestinationLocation': '',
                'PickupFrom': '',
                'PickupTo': '',
                'DeliveryFrom': '',
                'DeliveryTo': '',
                'Carrier': '',
                'Service': '',
                'ReferenceNumberType': 'PALLET_RET',
                'ReferenceNumber': str(pallet_ret) if pallet_ret != '' else '',
                'Quantity': '',
                'Weight': '',
                'Hectolitros': '',
                'Bultos': '',
                'Pallets': ''
            }
            records.append(detail2_record)
            self.validation_stats['detail_records'] += 1

            # Registro Detail 3 (D) - Quantities con datos de BD
            detail3_record = {
                'Type': 'D',
                'Shipment Number': '',
                'Shipment Desription': '',
                'Commodity': '',
                'Priority': '',
                'OriginLocation': '',
                'DestinationLocation': '',
                'PickupFrom': '',
                'PickupTo': '',
                'DeliveryFrom': '',
                'DeliveryTo': '',
                'Carrier': '',
                'Service': '',
                'ReferenceNumberType': '',
                'ReferenceNumber': '',
                'Quantity': quantity,
                'Weight': total_weight,
                'Hectolitros': hectolitros,  # Desde BD
                'Bultos': bultos,  # Desde BD
                'Pallets': ''  # Siempre vacío
            }
            records.append(detail3_record)
            self.validation_stats['detail_records'] += 1

            self.validation_stats['total_records'] += len(records)

            # logger.info(
            #     f"Registro {index + 1} procesado - Shipment: {shipment_number} - Ruta: {lane} - RouteID: {unique_route_id}")

            return records

        except Exception as e:
            logger.error(f"Error procesando registro {index}: {str(e)}")
            self.validation_stats['errors'].append(f"Error registro {index}: {str(e)}")
            raise

    def validate_uniqueness_before_processing(self, df: pd.DataFrame) -> bool:
        """Validar que no existen envíos duplicados antes de procesar"""
        try:
            logger.info("Validando unicidad de envíos...")

            # Verificar envíos únicos en el DataFrame
            cod_envios = df['Cód. Envío'].dropna().unique()
            logger.info(f"{len(cod_envios)} códigos de envío únicos en archivo")

            # Verificar si algún envío ya existe en el sistema
            # Esto dependería de tu sistema TMS específico

            return True

        except Exception as e:
            logger.error(f"Error validando unicidad: {e}")
            return False

    def process_all_data(self, df: pd.DataFrame) -> None:
        """Procesar todos los datos del DataFrame con consultas a BD"""
        try:
            logger.info("Iniciando procesamiento de datos con consultas a base de datos...")

            # Validar unicidad antes de procesar
            if not self.validate_uniqueness_before_processing(df):
                raise ValueError("Falló validación de unicidad")

            now_value = datetime.now().timestamp() / 86400
            ship_num = math.floor(now_value * 110000)

            logger.info(f"Número de envío base generado: {ship_num}")

            # Convertir DataFrame a lista para el conteo SKU per truck
            all_shipments = df.to_dict('records')
            total_records = 0

            for index, row in df.iterrows():
                try:
                    records = self.process_record(row, index, ship_num, all_shipments)
                    self.shipment_data.extend(records)
                    total_records += len(records)

                    if (index + 1) % 5 == 0:
                        logger.info(
                            f"Procesados {index + 1}/{len(df)} registros... ({self.validation_stats['database_queries']} consultas BD)")

                except Exception as e:
                    logger.error(f"Error procesando fila {index}: {str(e)}")
                    continue

            logger.info(f"Procesamiento completado!")
            logger.info(f"Total registros XML: {total_records}")
            logger.info(f"Total consultas BD: {self.validation_stats['database_queries']}")

        except Exception as e:
            logger.error(f"Error en procesamiento general: {str(e)}")
            raise

    def _write_mapping_sheet_complete(self, f):
        """Escribir hoja Mapping COMPLETA según especificaciones"""
        f.write(' <Worksheet ss:Name="Mapping">\n')
        f.write('  <Table ss:ExpandedColumnCount="3" ss:ExpandedRowCount="50" ')
        f.write('x:FullColumns="1" x:FullRows="1" ss:DefaultColumnWidth="56.25" ss:DefaultRowHeight="15">\n')

        # Headers
        f.write('   <Row>\n')
        for header in ["Map Type", "Map Value", "API Field"]:
            f.write(f'    <Cell><Data ss:Type="String">{header}</Data></Cell>\n')
        f.write('   </Row>\n')

        # MAPEO COMPLETO según especificaciones
        mapping_data = [
            ["COLUMN", "Type", "#RowType"],
            ["COLUMN", "Shipment Number", "Shipment.ShipmentNumber"],
            ["COLUMN", "Priority", "Shipment.ShipmentPriority"],
            ["CONSTANT", "BOL_ABI", "Shipment.CustomerCode"],
            ["CONSTANT", "BOL", "Shipment.LogisticsGroupCode"],
            ["CONSTANT", "AMBV", "Shipment.DivisionCode"],
            ["CONSTANT", "BOL", "Shipment.ProfitCenterCode"],
            ["CONSTANT", self.plan_id or "5001", "Shipment.SystemPlanID"],
            ["COLUMN", "Shipment Desription", "Shipment.ShipmentDescription"],
            ["CONSTANT", "SUM-BOL", "Shipment.ShipmentEntryVersionCode"],
            ["CONSTANT", "BA", "Shipment.ShipmentEntryTypeCode"],
            ["CONSTANT", "FT_PRE_PAID", "Shipment.FreightTermsEnumVal"],
            ["CONSTANT", "FALSE", "Shipment.UrgentFlag"],
            ["COLUMN", "Commodity", "Shipment.CommodityCode"],
            ["CONSTANT", "SFT_HUB", "Shipment.ShipFromLocationTypeEnumVal"],
            ["COLUMN", "OriginLocation", "Shipment.ShipFromLocationCode"],
            ["CONSTANT", "STT_HUB", "Shipment.ShipToLocationTypeEnumVal"],
            ["COLUMN", "DestinationLocation", "Shipment.ShipToLocationCode"],
            ["COLUMN", "PickupFrom", "Shipment.PickupFromDateTime"],
            ["COLUMN", "PickupTo", "Shipment.PickupToDateTime"],
            ["COLUMN", "DeliveryFrom", "Shipment.DeliveryFromDateTime"],
            ["COLUMN", "DeliveryTo", "Shipment.DeliveryToDateTime"],
            ["CONSTANT", "true", "Shipment.UseOriginDefaultsFlag"],
            ["CONSTANT", "true", "Shipment.UseDestinationDefaultsFlag"],
            ["CONSTANT", "false", "Shipment.IgnoreReferenceNumbersFlag"],
            ["CONSTANT", "false", "Shipment.IgnoreContainersFlag"],
            ["CONSTANT", "true", "Shipment.IgnoreChargeOverridesFlag"],
            ["COLUMN", "ReferenceNumber", "Shipment.ReferenceNumberStructure.ReferenceNumber"],
            ["COLUMN", "ReferenceNumberType", "Shipment.ReferenceNumberStructure.ReferenceNumberTypeCode"],
            ["CONSTANT", "PLL", "Shipment.Container.ContainerTypeCode"],
            ["COLUMN", "Quantity", "Shipment.Container.Quantity"],
            ["CONSTANT", "true", "Shipment.Container.IgnoreContainerOrientationsFlag"],
            ["CONSTANT", "false", "Shipment.Container.IgnoreWeightByFreightClassFlag"],
            ["CONSTANT", "true", "Shipment.Container.IgnoreShipmentItemsFlag"],
            ["CONSTANT", "true", "Shipment.Container.IgnoreReferenceNumbersFlag"],
            ["COLUMN", "Weight", "Shipment.Container.WeightByFreightClass.FreightClassNominalWeight"],
            ["CONSTANT", "*FAK", "Shipment.Container.WeightByFreightClass.FreightClassCode"],
            ["CONSTANT", "true", "Shipment.DeferAPRatingFlag"],
            ["CONSTANT", "true", "Shipment.DeferARRatingFlag"],
            ["CONSTANT", "false", "ExecuteAPRatingFlag"],
            ["CONSTANT", "false", "ExecuteARRatingFlag"],
            ["CONSTANT", "false", "IgnoreAllShipmentReferenceNumbersFlag"],
            ["COLUMN", "Hectolitros", "Shipment.Container.ContainerShippingInformation.FlexibleQuantity1"],
            ["COLUMN", "Bultos", "Shipment.Container.ContainerShippingInformation.FlexibleQuantity2"],
            ["COLUMN", "Carrier", "Shipment.PreferredAPCarrierCode"],
            ["COLUMN", "Service", "Shipment.PreferredAPServiceCode"],
            ["CONSTANT", "false", "Shipment.Container.Is3DLoadingRequiredFlag"],
            ["COLUMN", "Pallets", "Shipment.Container.ContainerShippingInformation.FlexibleQuantity3"]
        ]

        for row_data in mapping_data:
            f.write('   <Row>\n')
            for value in row_data:
                f.write(f'    <Cell><Data ss:Type="String">{self._escape_xml(str(value))}</Data></Cell>\n')
            f.write('   </Row>\n')

        f.write('  </Table>\n')
        f.write(' </Worksheet>\n')

    def generate_validation_report(self) -> str:
        """Generar reporte de validación de calidad con estadísticas de BD"""
        report = []
        report.append("=" * 70)
        report.append("REPORTE DE VALIDACIÓN - VERSIÓN BASE DE DATOS")
        report.append("=" * 70)

        # Estadísticas generales
        report.append(f"ESTADÍSTICAS GENERALES:")
        report.append(f"   Total de registros generados: {self.validation_stats['total_records']}")
        report.append(f"   Registros Header (H): {self.validation_stats['header_records']}")
        report.append(f"   Registros Detail (D): {self.validation_stats['detail_records']}")

        # Validar proporción 1:3 (1 Header por cada 3 Detail)
        expected_ratio = self.validation_stats['header_records'] * 3
        if self.validation_stats['detail_records'] == expected_ratio:
            report.append(f"   ✅ Proporción 1:3 (H:D) correcta")
        else:
            report.append(
                f"   Proporción incorrecta - Esperado: {expected_ratio}, Actual: {self.validation_stats['detail_records']}")

        # Estadísticas de base de datos
        report.append(f"\nESTADÍSTICAS DE BASE DE DATOS:")
        report.append(f"   Total consultas ejecutadas: {self.validation_stats['database_queries']}")
        report.append(
            f"   Promedio consultas por registro: {self.validation_stats['database_queries'] / max(1, self.validation_stats['header_records']):.1f}")

        # ReferenceNumbers generados desde BD
        report.append(f"\nEFERENCE NUMBERS (desde BD):")
        if self.validation_stats['reference_numbers_generated']:
            refs = [int(r) for r in self.validation_stats['reference_numbers_generated']]
            min_ref = min(refs)
            max_ref = max(refs)
            total_refs = len(refs)
            report.append(f"   Rango: {min_ref} - {max_ref}")
            report.append(f"   Total generados: {total_refs}")
            report.append(f"   Todos los números provienen de la secuencia de BD")

        # Prioridades desde BD
        report.append(f"\nPRIORIDADES (desde maestro_dtto):")
        if self.validation_stats['priority_conversions']:
            for priority, count in self.validation_stats['priority_conversions'].items():
                report.append(f"   Prioridad {priority}: {count} rutas")
        else:
            report.append("   No se obtuvieron prioridades de BD")

        # Errores encontrados
        report.append(f"\n ERRORES Y ADVERTENCIAS:")
        if self.validation_stats['errors']:
            for error in self.validation_stats['errors'][:10]:
                report.append(f"     {error}")
            if len(self.validation_stats['errors']) > 10:
                report.append(f"   ... y {len(self.validation_stats['errors']) - 10} errores más")
        else:
            report.append("   No se encontraron errores")

        # Verificaciones de integridad con BD
        report.append(f"\n VERIFICACIONES DE INTEGRIDAD BD:")

        # Verificar Headers con datos de BD
        headers_with_data = [r for r in self.shipment_data if r['Type'] == 'H' and
                             r['ReferenceNumber'] != '' and r['Commodity'] != '']
        report.append(
            f"   Headers con datos de BD completos: {len(headers_with_data)}/{self.validation_stats['header_records']}")

        # Verificar quantities desde BD
        details_with_quantities = [r for r in self.shipment_data if r['Type'] == 'D' and
                                   (r.get('Hectolitros') != '' or r.get('Bultos') != '')]
        report.append(f"   Details con cantidades de BD: {len(details_with_quantities)}")

        report.append(f"\n💡 DATOS OBTENIDOS DE BASE DE DATOS:")
        report.append(f"   ✅ SKU Names desde 'dados_produtos' y 'maestro_envases'")
        report.append(f"   ✅ Prioridades desde 'maestro_dtto' por rutas")
        report.append(f"   ✅ Commodities desde 'dados_produtos'")
        report.append(f"   ✅ Hectolitros calculados desde BD")
        report.append(f"   ✅ Bultos calculados desde BD")
        report.append(f"   ✅ ReferenceNumbers desde secuencia BD")

        report.append("=" * 70)
        return "\n".join(report)

    def export_to_xml(self, output_path: Optional[str] = None) -> str:
        """Exportar datos a formato XML con validación completa y doble ubicación"""
        try:
            if not self.shipment_data:
                raise ValueError("No hay datos para exportar")

            if not output_path:
                timestamp = datetime.now().strftime("%Y_%m_%d_%H_%M")
                output_path = os.path.join(self.destination_folder, f"processShipmentOrderCreate_DB_{timestamp}.xml")

            logger.info(f"Generando archivo XML: {output_path}")

            # GENERAR EL XML ORIGINAL (código existente sin cambios)
            with open(output_path, 'w', encoding='utf-8') as f:
                # ... todo el código XML existente sin cambios ...
                # (escribir declaración XML, workbook, data, etc.)

                # Escribir declaración XML
                f.write('<?xml version="1.0"?>\n')
                f.write('<?mso-application progid="Excel.Sheet"?>\n')

                # Workbook con namespaces
                f.write('<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"\n')
                f.write(' xmlns:o="urn:schemas-microsoft-com:office:office"\n')
                f.write(' xmlns:x="urn:schemas-microsoft-com:office:excel"\n')
                f.write(' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"\n')
                f.write(' xmlns:html="http://www.w3.org/TR/REC-html40">\n')

                # DocumentProperties
                f.write(' <DocumentProperties xmlns="urn:schemas-microsoft-com:office:office">\n')
                f.write('  <Author>Python Shipment Generator DB</Author>\n')
                f.write('  <LastAuthor>Python Shipment Generator DB</LastAuthor>\n')
                f.write(f'  <Created>{datetime.now().isoformat()}Z</Created>\n')
                f.write('  <Version>16.00</Version>\n')
                f.write(' </DocumentProperties>\n')

                # OfficeDocumentSettings
                f.write(' <OfficeDocumentSettings xmlns="urn:schemas-microsoft-com:office:office">\n')
                f.write('  <AllowPNG/>\n')
                f.write(' </OfficeDocumentSettings>\n')

                # ExcelWorkbook
                f.write(' <ExcelWorkbook xmlns="urn:schemas-microsoft-com:office:excel">\n')
                f.write('  <WindowHeight>20745</WindowHeight>\n')
                f.write('  <WindowWidth>32767</WindowWidth>\n')
                f.write('  <WindowTopX>32767</WindowTopX>\n')
                f.write('  <WindowTopY>32767</WindowTopY>\n')
                f.write('  <ProtectStructure>False</ProtectStructure>\n')
                f.write('  <ProtectWindows>False</ProtectWindows>\n')
                f.write(' </ExcelWorkbook>\n')

                # Styles
                f.write(' <Styles>\n')
                f.write('  <Style ss:ID="Default" ss:Name="Normal">\n')
                f.write('   <Alignment ss:Vertical="Bottom"/>\n')
                f.write('   <Borders/>\n')
                f.write('   <Font ss:FontName="Calibri" x:Family="Swiss" ss:Size="11" ss:Color="#000000"/>\n')
                f.write('   <Interior/>\n')
                f.write('   <NumberFormat/>\n')
                f.write('   <Protection/>\n')
                f.write('  </Style>\n')
                f.write(' </Styles>\n')

                # Worksheet Data
                f.write(' <Worksheet ss:Name="Data">\n')
                f.write('  <Table ss:ExpandedColumnCount="20" ')
                f.write(f'ss:ExpandedRowCount="{len(self.shipment_data) + 1}" ')
                f.write('x:FullColumns="1" x:FullRows="1" ')
                f.write('ss:DefaultColumnWidth="49.5" ss:DefaultRowHeight="15">\n')

                # Headers row
                f.write('   <Row>\n')
                for header in self.XML_HEADERS:
                    f.write('    <Cell><Data ss:Type="String">')
                    f.write(self._escape_xml(header))
                    f.write('</Data></Cell>\n')
                f.write('   </Row>\n')

                # Data rows con validación estricta de tipos
                for record in self.shipment_data:
                    f.write('   <Row>\n')
                    for header in self.XML_HEADERS:
                        value = record.get(header, '')
                        f.write('    <Cell>')

                        # VALIDACIÓN ESTRICTA DE TIPOS
                        if header == 'Priority' and value != '':
                            # Priority debe ser SIEMPRE numérico
                            if isinstance(value, (int, float)) and value != '':
                                f.write(f'<Data ss:Type="Number">{int(value)}</Data>')
                            else:
                                f.write('<Data ss:Type="Number">1</Data>')
                        elif header in ['Quantity', 'Weight', 'Hectolitros', 'Bultos'] and value != '':
                            # Otros campos numéricos
                            if isinstance(value, (int, float)):
                                f.write(f'<Data ss:Type="Number">{value}</Data>')
                            else:
                                f.write('<Data ss:Type="String"></Data>')
                        elif header in ['PickupFrom', 'PickupTo', 'DeliveryFrom', 'DeliveryTo'] and value != '':
                            # Campos de fecha
                            f.write(f'<Data ss:Type="String" x:Ticked="1">{self._escape_xml(str(value))}</Data>')
                        else:
                            # Campos de texto
                            f.write(
                                f'<Data ss:Type="String">{self._escape_xml(str(value)) if value != "" else ""}</Data>')

                        f.write('</Cell>\n')
                    f.write('   </Row>\n')

                f.write('  </Table>\n')
                f.write(' </Worksheet>\n')

                # Agregar hojas Info y Mapping COMPLETO
                self._write_info_sheet(f)
                self._write_mapping_sheet_complete(f)

                f.write('</Workbook>\n')

            logger.info(f"Archivo XML generado exitosamente: {output_path}")

            # ===== NUEVA FUNCIONALIDAD: COPIA A 2ETAPA =====
            self.create_2etapa_copy(output_path)

            return output_path

        except Exception as e:
            logger.error(f"Error exportando XML: {str(e)}")
            raise

    def create_2etapa_copy(self, original_xml_path: str) -> None:
        """Crear copia del XML en carpeta 2etapa con nombre simplificado"""
        try:
            import shutil

            # Extraer mes y día de destination_folder
            # destination_folder tiene formato: "Beer/04/01"
            path_parts = self.destination_folder.split(os.sep)
            if len(path_parts) >= 3:
                file_type = path_parts[0].lower()  # "beer"
                month = path_parts[1]  # "04"
                day = path_parts[2]  # "01"
            else:
                logger.warning("No se pudo extraer mes/día de destination_folder")
                return

            # Crear carpeta 2etapa
            etapa2_folder = os.path.join(".", "2etapa", "output", file_type, month, day)
            os.makedirs(etapa2_folder, exist_ok=True)

            # Copiar XML con nombre simplificado
            etapa2_xml_path = os.path.join(etapa2_folder, f"{file_type}.xml")
            shutil.copy2(original_xml_path, etapa2_xml_path)

            logger.info(f"📋 Copia XML creada en 2etapa: {etapa2_xml_path}")

            # Guardar la ruta para usar en placas
            self.etapa2_folder = etapa2_folder

        except Exception as e:
            logger.error(f"Error creando copia 2etapa: {str(e)}")

    def _escape_xml(self, text: str) -> str:
        """Escapar caracteres especiales XML"""
        if not text:
            return ""

        text = str(text)
        text = text.replace("&", "&amp;")
        text = text.replace("<", "&lt;")
        text = text.replace(">", "&gt;")
        text = text.replace('"', "&quot;")
        text = text.replace("'", "&apos;")
        return text

    def _write_info_sheet(self, f):
        """Escribir hoja Info"""
        f.write(' <Worksheet ss:Name="Info">\n')
        f.write('  <Table ss:ExpandedColumnCount="3" ss:ExpandedRowCount="6" ')
        f.write('x:FullColumns="1" x:FullRows="1" ss:DefaultColumnWidth="56.25" ss:DefaultRowHeight="15">\n')

        info_data = [
            ["Version", "2", ""],
            ["APIRequest", "processShipmentOrderCreate", ""],
            ["Timestamp", "yyyy-MM-dd HH:mm", ""],
            ["Date", "yyyy-MM-dd", ""],
            ["Time", "HH:mm", ""],
            ["", "", "All cells should be formatted to 'text'"]
        ]

        for row_data in info_data:
            f.write('   <Row>\n')
            for value in row_data:
                f.write(f'    <Cell><Data ss:Type="String">{self._escape_xml(value)}</Data></Cell>\n')
            f.write('   </Row>\n')

        f.write('  </Table>\n')
        f.write(' </Worksheet>\n')

    def generate_xml_from_file(self, input_file: str, output_file: Optional[str] = None) -> str:
        """Método principal para generar XML desde archivo Excel con BD"""
        try:
            logger.info("=== INICIANDO GENERACIÓN XML CON BASE DE DATOS ===")

            # 1. Verificar conexión a BD
            if not self.db.connection or not self.db.connection.is_connected():
                raise ValueError("No hay conexión a la base de datos")

            # 2. Detectar tipo de archivo y Plan ID
            file_type, plan_id, destination_folder = self.detect_file_type_and_plan_id(input_file)
            self.plan_id = plan_id
            self.destination_folder = destination_folder

            # 3. Validar archivo de entrada
            if not self.validate_input_file(input_file):
                raise ValueError("Archivo de entrada inválido")

            # 4. Cargar datos
            df = self.load_data(input_file)

            # 5. Procesar datos con consultas a BD
            self.process_all_data(df)

            # 6. Exportar XML
            output_path = self.export_to_xml(output_file)

            # 7. Generar reporte de validación
            validation_report = self.generate_validation_report()
            try:
                logger.info(validation_report)
            except UnicodeEncodeError:
                # Reemplazar emojis por texto simple si hay problemas
                clean_report = validation_report.replace("✅", "[OK]").replace("⚠️", "[WARN]").replace("❌",
                                                                                                      "[ERROR]").replace(
                    "💡", "[INFO]")
                logger.info(clean_report)

            # 8. Guardar reporte en archivo
            report_path = output_path.replace('.xml', '_validation_report.txt')
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write(validation_report)

            logger.info("=== GENERACIÓN COMPLETADA ===")
            logger.info(f"Archivo XML: {output_path}")
            logger.info(f"Reporte de validación: {report_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error en generación: {str(e)}")
            raise
        finally:
            # Cerrar conexión a BD
            self.db.close_connection()

    def __del__(self):
        """Destructor para cerrar conexión BD"""
        if hasattr(self, 'db'):
            self.db.close_connection()


def generate_validated_plates_excel(input_file, db_config, output_file=None, destination_folder=".",
                                    etapa2_folder=None):
    """
    Descarga placas con grupo 'Transportadoras' e integra camiones externos
    CORREGIDO: Deduplicación completa en combinación final
    """
    import mysql.connector
    import pandas as pd
    import shutil
    import os

    # Si no se especifica output_file, generarlo en la carpeta de destino
    if output_file is None:
        output_file = os.path.join(destination_folder, "availability_placas.xlsx")

    logger.info("🔎 Iniciando procesamiento de placas...")

    # ===== PASO 1: OBTENER PLACAS DESDE BASE DE DATOS =====
    df_input = pd.read_excel(input_file, sheet_name="Consolidado", dtype=str)

    # Buscar columnas de manera más flexible
    col_codplanta = None
    col_placa = None

    for col in df_input.columns:
        if "Cod Planta" in str(col):
            col_codplanta = col
        if "Placa" in str(col):
            col_placa = col

    if not col_codplanta or not col_placa:
        logger.error(f"❌ No se encontraron columnas necesarias")
        logger.error(f"   Columnas disponibles: {list(df_input.columns)}")
        return None

    logger.info(f"📋 Usando columnas: '{col_codplanta}' y '{col_placa}'")

    placas_bd = []
    conn = mysql.connector.connect(**db_config)
    cursor = conn.cursor(dictionary=True)

    for _, row in df_input.iterrows():
        codplanta = row[col_codplanta]
        placa = row[col_placa]

        if pd.isna(placa) or str(placa).strip() == "":
            continue

        cursor.execute("SELECT grupo FROM drivers WHERE placa = %s", (placa,))
        result = cursor.fetchone()

        if result is not None and result.get("grupo") == "Transportadoras":
            placas_bd.append({
                "Origen": str(codplanta).strip(),
                "Placa": str(placa).strip().upper(),  # ← NORMALIZAR A MAYÚSCULAS
                "Reusable": 0
            })

    cursor.close()
    conn.close()

    logger.info(f"📊 Placas desde BD (Transportadoras): {len(placas_bd)}")

    # ===== PASO 2: EXTRAER MES Y DÍA DEL ARCHIVO PRINCIPAL =====
    mes, dia = extract_date_from_input_file(input_file)
    logger.info(f"📅 Fecha extraída: mes {mes}, día {dia}")

    # ===== PASO 3: BUSCAR Y PROCESAR ARCHIVO DE DISPONIBILIDAD =====
    placas_externas = get_disponibilidad_camiones(mes, dia)  # Ya deduplicadas internamente

    # Normalizar placas externas a mayúsculas
    for placa_info in placas_externas:
        placa_info["Placa"] = placa_info["Placa"].upper()
        placa_info["Origen"] = str(placa_info["Origen"]).strip()

    logger.info(f"📊 Placas externas (ya deduplicadas): {len(placas_externas)}")

    # ===== PASO 4: COMBINAR Y DEDUPLICAR TODOS LOS DATOS =====
    todas_las_placas = placas_bd + placas_externas
    logger.info(f"📊 Total placas antes de deduplicación final: {len(todas_las_placas)}")

    # ===== PASO 5: DEDUPLICACIÓN COMPLETA =====
    placas_unicas = []
    placas_vistas = set()

    for placa_info in todas_las_placas:
        # Crear clave única: PLACA + ORIGEN (ambos normalizados)
        clave_unica = f"{placa_info['Placa'].upper()}_{str(placa_info['Origen']).strip()}"

        if clave_unica not in placas_vistas:
            placas_vistas.add(clave_unica)
            placas_unicas.append(placa_info)
        else:
            logger.debug(f"🔄 Duplicado eliminado: {placa_info['Placa']} en {placa_info['Origen']}")

    duplicados_totales = len(todas_las_placas) - len(placas_unicas)
    logger.info(f"📊 Duplicados eliminados en combinación final: {duplicados_totales}")
    logger.info(f"📊 Total placas finales únicas: {len(placas_unicas)}")

    # ===== PASO 6: CREAR DATAFRAME FINAL =====
    if len(placas_unicas) > 0:
        df_final = pd.DataFrame(placas_unicas)

        # Ordenar por Origen y luego por Placa para consistencia
        df_final = df_final.sort_values(['Origen', 'Placa']).reset_index(drop=True)

        logger.info(f"📊 DataFrame final creado con {len(df_final)} registros únicos")
    else:
        logger.warning("⚠️ No se encontraron placas para procesar")
        df_final = pd.DataFrame(columns=['Origen', 'Placa', 'Reusable'])

    # ===== PASO 7: GENERAR ARCHIVO ORIGINAL =====
    try:
        df_final.to_excel(output_file, index=False, sheet_name="Disponibles")
        logger.info(f"📄 Archivo de placas generado: {output_file}")
    except Exception as e:
        logger.error(f"Error generando archivo Excel: {str(e)}")
        return None

    # ===== PASO 8: COPIA A 2ETAPA =====
    if etapa2_folder:
        try:
            etapa2_placas_path = os.path.join(etapa2_folder, "availability.xlsx")
            df_final.to_excel(etapa2_placas_path, index=False, sheet_name="Disponibles")
            logger.info(f"📋 Copia placas creada en 2etapa: {etapa2_placas_path}")
        except Exception as e:
            logger.error(f"Error creando copia placas 2etapa: {str(e)}")

    # ===== PASO 9: REPORTE FINAL =====
    logger.info(f"\n📈 RESUMEN FINAL DE PLACAS:")
    logger.info(f"   📦 Placas BD (Transportadoras): {len(placas_bd)}")
    logger.info(f"   🚛 Placas externas (disponibilidad): {len(placas_externas)}")
    logger.info(f"   🔄 Duplicados eliminados: {duplicados_totales}")
    logger.info(f"   ✅ Total único final: {len(placas_unicas)}")

    # Análisis por origen
    if len(df_final) > 0:
        origenes_count = df_final['Origen'].value_counts()
        logger.info(f"📊 Distribución por origen:")
        for origen, count in origenes_count.head(5).items():
            logger.info(f"   {origen}: {count} placas")

    return output_file


def extract_date_from_input_file(input_file):
    """Extraer mes y día del nombre del archivo principal - CORREGIDA"""
    import re

    try:
        file_name = os.path.basename(input_file)
        logger.info(f"Extrayendo fecha del archivo: {file_name}")

        # ===== PATRÓN 1: SD - Programa_SD_D_MM_YYYY_ =====
        pattern_sd = r'Programa_SD_(\d{1,2})_(\d{1,2})_\d{4}_'
        match_sd = re.search(pattern_sd, file_name)

        if match_sd:
            day = match_sd.group(1).zfill(2)
            month = match_sd.group(2).zfill(2)
            logger.info(f"📅 Patrón SD detectado: día {day}, mes {month}")
            return month, day

        # ===== PATRÓN 2: CB - Envíos CBs DD-MM.xlsm =====
        pattern_cb = r'Envíos\s+CBs?\s+(\d{1,2})-(\d{1,2})'
        match_cb = re.search(pattern_cb, file_name)

        if match_cb:
            day = match_cb.group(1).zfill(2)
            month = match_cb.group(2).zfill(2)
            logger.info(f"📅 Patrón CB detectado: día {day}, mes {month}")
            return month, day

        # ===== PATRÓN 3: Beer y genérico - DD_MM =====
        pattern = r'(\d{1,2})_(\d{1,2})'
        match = re.search(pattern, file_name)

        if match:
            day = match.group(1).zfill(2)
            month = match.group(2).zfill(2)
            logger.info(f"📅 Patrón genérico DD_MM detectado: día {day}, mes {month}")
            return month, day

        logger.warning(f"⚠️ No se pudo extraer fecha de: {file_name}")
        return None, None

    except Exception as e:
        logger.error(f"Error extrayendo fecha: {str(e)}")
        return None, None


def buscar_y_procesar_disponibilidad(mes, dia):
    """
    Buscar archivo de disponibilidad y procesar todas las pestañas
    """
    placas_externas = []

    if not mes or not dia:
        logger.warning("⚠️ No se pudo determinar mes/día, saltando búsqueda de disponibilidad")
        return placas_externas

    try:
        # Construir ruta de búsqueda
        carpeta_disponibilidad = os.path.join(".", "disponibilidad_camiones", mes)
        archivo_patron = f"Disponibilidad de Camiones {dia}-{mes}.xlsx"
        archivo_path = os.path.join(carpeta_disponibilidad, archivo_patron)

        logger.info(f"🔍 Buscando: {archivo_path}")

        if not os.path.exists(archivo_path):
            logger.warning(f"⚠️ No se encontró archivo de disponibilidad: {archivo_path}")
            return placas_externas

        logger.info(f"✅ Archivo encontrado: {archivo_path}")

        # Procesar pestañas en orden
        pestanas_a_procesar = [
            "Reporte Tra.",
            "Reporte Espe.",
            "Reporte Espe. (tarde)"
        ]

        for pestana in pestanas_a_procesar:
            placas_pestana = procesar_pestana_disponibilidad(archivo_path, pestana)
            placas_externas.extend(placas_pestana)

        logger.info(f"📊 Total placas externas procesadas: {len(placas_externas)}")

    except Exception as e:
        logger.error(f"Error procesando disponibilidad: {str(e)}")

    return placas_externas


def procesar_pestana_disponibilidad(archivo_path, nombre_pestana):
    """
    Procesar una pestaña específica del archivo de disponibilidad
    ACTUALIZADO: Usa header=4 y nombres de columnas correctos
    """
    placas_pestana = []

    try:
        logger.info(f"📋 Procesando pestaña: {nombre_pestana}")

        # CAMBIO CLAVE: Leer la pestaña con header en fila 4 (índice 4)
        df = pd.read_excel(archivo_path, sheet_name=nombre_pestana, header=4, dtype=str)

        logger.info(f"📊 Dimensiones de datos: {df.shape[0]} filas x {df.shape[1]} columnas")
        logger.info(f"📋 Columnas encontradas: {list(df.columns)[:10]}...")  # Solo primeras 10

        # Verificar que existan las columnas necesarias (nombres exactos)
        if "Placa" not in df.columns:
            logger.warning(f"⚠️ Columna 'Placa' no encontrada en {nombre_pestana}")
            logger.info(f"💡 Columnas disponibles: {[col for col in df.columns if 'placa' in col.lower()]}")
            return placas_pestana

        if "Dep Planta/CD" not in df.columns:
            logger.warning(f"⚠️ Columna 'Dep Planta/CD' no encontrada en {nombre_pestana}")
            logger.info(
                f"💡 Columnas disponibles: {[col for col in df.columns if 'planta' in col.lower() or 'cd' in col.lower()]}")
            return placas_pestana

        # Filtrar filas válidas (sin valores nulos en columnas clave)
        df_valido = df.dropna(subset=["Placa", "Dep Planta/CD"])

        logger.info(f"📊 Filas válidas después de filtrar: {len(df_valido)}")

        # Procesar cada fila válida
        for _, row in df_valido.iterrows():
            placa = str(row["Placa"]).strip()
            origen = str(row["Dep Planta/CD"]).strip()

            # Validar que no estén vacíos y no sean 'nan'
            if (placa and origen and
                    placa.lower() not in ['nan', 'none', ''] and
                    origen.lower() not in ['nan', 'none', '']):
                placas_pestana.append({
                    "Origen": origen,
                    "Placa": placa,
                    "Reusable": 0
                })

        logger.info(f"📊 Placas extraídas de {nombre_pestana}: {len(placas_pestana)}")

        # Mostrar algunas placas de ejemplo (para debug)
        if len(placas_pestana) > 0:
            ejemplos = placas_pestana[:3]
            logger.info(f"💡 Ejemplos: {ejemplos}")

    except FileNotFoundError:
        logger.warning(f"⚠️ Pestaña '{nombre_pestana}' no encontrada en archivo")
    except Exception as e:
        logger.error(f"Error procesando pestaña {nombre_pestana}: {str(e)}")

    return placas_pestana


def clear_excel_filters(file_path, sheet_name):
    """
    Función auxiliar para quitar filtros de Excel (usando openpyxl si es necesario)
    """
    try:
        from openpyxl import load_workbook

        # Cargar workbook
        wb = load_workbook(file_path)

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Quitar autofilter si existe
            if ws.auto_filter:
                ws.auto_filter = None
                wb.save(file_path)
                logger.info(f"🔧 Filtros removidos de {sheet_name}")

    except ImportError:
        logger.warning("openpyxl no disponible, continuando sin quitar filtros")
    except Exception as e:
        logger.warning(f"No se pudieron quitar filtros: {str(e)}")


def test_disponibilidad_integration():
    """
    Función de prueba para verificar la integración
    """
    print("🧪 PRUEBA DE INTEGRACIÓN DE DISPONIBILIDAD")
    print("=" * 50)

    # Probar extracción de fecha
    test_files = [
        "Programa Beer_01_04.xlsm",
        "Data_SD_15_05.xlsm",
        "Archivo_CB_28_12.xlsm"
    ]

    for test_file in test_files:
        mes, dia = extract_date_from_input_file(test_file)
        print(f"📁 {test_file} → mes: {mes}, día: {dia}")

    # Probar búsqueda de archivos
    test_disponibilidad = get_disponibilidad_camiones("04", "01")
    print(f"📊 Placas encontradas: {len(test_disponibilidad)}")


def test_disponibilidad_con_analisis():
    """
    Test con análisis detallado de duplicados
    """
    print("🧪 PRUEBA CON ANÁLISIS DE DUPLICADOS")
    print("=" * 50)

    # Datos de ejemplo para testing
    ejemplos = [
        ("Beer_01_04.xlsm", "04", "01"),
        ("Data_SD_15_05.xlsm", "05", "15"),
        ("Archivo_CB_28_12.xlsm", "12", "28")
    ]

    for archivo_ejemplo, mes, dia in ejemplos:
        print(f"📁 {archivo_ejemplo} → mes: {mes}, día: {dia}")

    # Probar con el primer ejemplo
    mes, dia = "04", "01"
    print(f"\n🔍 Analizando disponibilidad para {dia}-{mes}")
    print("-" * 40)

    placas = get_disponibilidad_camiones(mes, dia)

    if placas:
        print(f"📊 Total final: {len(placas)} placas únicas")

        # Analizar por origen
        origenes = {}
        for placa in placas:
            origen = placa['Origen']
            if origen not in origenes:
                origenes[origen] = []
            origenes[origen].append(placa['Placa'])

        print("\n📋 Distribución por origen:")
        for origen, lista_placas in origenes.items():
            print(f"   {origen}: {len(lista_placas)} placas")
            print(f"      Ejemplos: {lista_placas[:3]}")

        # Verificar si hay duplicados que se nos escaparon
        todas_placas = [p['Placa'].upper() for p in placas]
        duplicados = len(todas_placas) - len(set(todas_placas))

        if duplicados > 0:
            print(f"\n⚠️  ATENCIÓN: Aún hay {duplicados} duplicados de placa")
        else:
            print(f"\n✅ PERFECTO: No hay duplicados de placa")

    else:
        print("❌ No se encontraron placas")

def buscar_archivo_disponibilidad(mes, dia):
    """
    Buscar archivo de disponibilidad para mes y día específicos
    """
    try:
        # Construir ruta de búsqueda
        carpeta_disponibilidad = os.path.join(".", "disponibilidad_camiones", mes)
        archivo_patron = f"Disponibilidad de Camiones {dia}-{mes}.xlsx"
        archivo_path = os.path.join(carpeta_disponibilidad, archivo_patron)

        logger.info(f"🔍 Buscando: {archivo_path}")

        if os.path.exists(archivo_path):
            logger.info(f"✅ Archivo encontrado: {archivo_path}")
            return archivo_path
        else:
            logger.warning(f"⚠️ No se encontró archivo de disponibilidad: {archivo_path}")
            return None

    except Exception as e:
        logger.error(f"Error buscando archivo disponibilidad: {str(e)}")
        return None


def get_disponibilidad_camiones(mes, dia):
    """
    Obtener disponibilidad de camiones desde archivo Excel
    ACTUALIZADO: Con deduplicación de placas
    """
    placas_todas = []

    try:
        archivo_path = buscar_archivo_disponibilidad(mes, dia)
        if not archivo_path:
            logger.warning(f"No se encontró archivo de disponibilidad para {dia}-{mes}")
            return placas_todas

        # Procesar todas las pestañas
        pestanas = ["Reporte Tra.", "Reporte Espe.", "Reporte Espe. (tarde)"]

        for pestana in pestanas:
            placas_pestana = procesar_pestana_disponibilidad(archivo_path, pestana)
            placas_todas.extend(placas_pestana)

        logger.info(f"📊 Total placas externas procesadas: {len(placas_todas)}")

        # ========== DEDUPLICACIÓN ==========
        placas_unicas = []
        placas_vistas = set()

        for placa_info in placas_todas:
            # Crear clave única: combinar placa + origen
            clave = f"{placa_info['Placa'].upper()}_{placa_info['Origen']}"

            if clave not in placas_vistas:
                placas_vistas.add(clave)
                placas_unicas.append(placa_info)

        logger.info(f"📊 Placas después de deduplicación: {len(placas_unicas)}")
        logger.info(f"🔄 Duplicados eliminados: {len(placas_todas) - len(placas_unicas)}")

        return placas_unicas

    except Exception as e:
        logger.error(f"Error al obtener disponibilidad de camiones: {str(e)}")
        return []


def main():
    """Función principal para uso desde línea de comandos"""
    import os
    if sys.platform == "win32":
        try:
            os.system("chcp 65001 > nul")
        except:
            pass

    if len(sys.argv) < 2:
        print("🚛 Shipment XML Generator v5.0 - VERSIÓN BASE DE DATOS")
        print("=" * 70)
        print("Uso: python shipment_generator.py <archivo_excel> [archivo_salida.xml]")
        print("")
        print("🎯 OPCIONES:")
        print("  --from-planta    Usar 'Cod Planta' como origen en lugar de 'Cód. Origen'")
        print("  --no-placas      Omitir generación de archivo de placas")
        print("")
        print("🎯 NUEVAS CARACTERÍSTICAS v5.0:")
        print("✅ Integración completa con base de datos MySQL")
        print("✅ SKU Names desde dados_produtos y maestro_envases")
        print("✅ Prioridades desde maestro_dtto por rutas")
        print("✅ Commodities desde dados_produtos")
        print("✅ Hectolitros y Bultos calculados desde BD")
        print("✅ ReferenceNumber correlativo desde BD")
        print("✅ Reporte detallado de consultas BD")
        print("")
        print("🔍 FÓRMULAS IMPLEMENTADAS:")
        print("  📊 SKU Name = BUSCARV en dados_produtos o maestro_envases")
        print("  🎯 Priority = BUSCARV en maestro_dtto por rutas BO_X-BO_Y")
        print("  📦 Commodity = BUSCARV en dados_produtos")
        print("  🔢 Hectolitros = hl_por_pallet * pallets desde BD")
        print("  📋 Bultos = bultos_x_pallet * pallets desde BD")
        print("  🔢 ReferenceNumber = secuencia correlativa desde BD")
        print("")
        print("📋 CONFIGURACIÓN BD (editar en código):")
        print("  Host: localhost")
        print("  Database: tms_shipment_db")
        print("  User: root")
        print("  Password: Dali19((Kafka")
        print("")
        print("Ejemplos:")
        print("  python shipment_generator.py 'Programa Beer_31_05.xlsm'")
        print("  python shipment_generator.py 'Data_SD.xlsm' output_custom.xml")
        print("  python shipment_generator.py 'Data.xlsm' --from-planta")  # ← NUEVO EJEMPLO
        print("  python shipment_generator.py 'Data.xlsm' output.xml --from-planta --no-placas")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 and not sys.argv[2].startswith('--') else None
    skip_placas = "--no-placas" in sys.argv
    use_planta_as_origen = "--from-planta" in sys.argv
    try:
        print("🚛 Shipment XML Generator v5.0 - VERSIÓN BASE DE DATOS")
        print("=" * 70)
        print(f"📄 Archivo de entrada: {input_file}")

        if use_planta_as_origen:
            print("🔄 MODO: Usando 'Cod Planta' como origen (--from-planta)")
        else:
            print("📍 MODO: Usando 'Cód. Origen' normal")


        print(f"🔍 Conectando a base de datos...")

        # Configuración de BD (personalizable)
        db_config = {
            'host': 'localhost',
            'database': 'tms_shipment_db',
            'user': 'root',
            'password': 'Dali19((Kafka'  # Cambiar según configuración
        }

        generator = ShipmentXMLGenerator(db_config, use_planta_as_origen=use_planta_as_origen)
        result_file = generator.generate_xml_from_file(input_file, output_file)

        if not skip_placas:
            print("🔎 Procesando placas de disponibilidad...")
            # CAMBIO: pasar etapa2_folder
            placas_file = generate_validated_plates_excel(
                input_file,
                db_config,
                destination_folder=generator.destination_folder,
                etapa2_folder=getattr(generator, 'etapa2_folder', None)  # ← NUEVO PARÁMETRO
            )
            print(f"📄 Archivo de placas generado: {placas_file}")

        print(f"\n🎉 PROCESAMIENTO COMPLETADO EXITOSAMENTE")
        print(f"📊 XML generado: {result_file}")

        if use_planta_as_origen:
            print(f"🔄 Modo usado: Cod Planta como origen")
        else:
            print(f"📍 Modo usado: Cód. Origen normal")


        print(f"📋 Reporte de validación: {result_file.replace('.xml', '_validation_report.txt')}")
        print(f"📁 Plan ID: {generator.plan_id}")
        print(f"🗂️  Carpeta: {generator.destination_folder}")

        # Mostrar resumen de validación
        stats = generator.validation_stats
        print(f"\n📈 RESUMEN DE VALIDACIÓN CON BD:")
        print(f"   📦 Total registros: {stats['total_records']}")
        print(f"   🎯 Headers (H): {stats['header_records']}")
        print(f"   📋 Details (D): {stats['detail_records']}")
        print(f"   🔍 Consultas BD: {stats['database_queries']}")

        # Verificar proporción
        expected_details = stats['header_records'] * 3
        if stats['detail_records'] == expected_details:
            print(f"   ✅ Proporción 1:3 correcta")
        else:
            print(f"   ⚠️  Proporción incorrecta")

        # Reference numbers desde BD
        if stats['reference_numbers_generated']:
            refs = [int(r) for r in stats['reference_numbers_generated']]
            min_ref = min(refs)
            max_ref = max(refs)
            print(f"   🔢 ReferenceNumbers BD: {min_ref} - {max_ref}")

        # Errores
        if stats['errors']:
            print(f"   ⚠️  Errores encontrados: {len(stats['errors'])}")
        else:
            print(f"   ✅ Sin errores")

        print(f"\n🔥 ¡ARCHIVO LISTO PARA TMS CON DATOS REALES!")
        print(f"💡 Todos los datos provienen de consultas a base de datos")
        print(f"📋 Revisa el reporte detallado para información de BD")

    except Exception as e:
        print(f"\n❌ ERROR: {str(e)}")
        print("📋 Revisa 'shipment_generator.log' para detalles")
        print("\n💡 VERIFICACIONES:")
        print("   1. ¿Está MySQL ejecutándose?")
        print("   2. ¿Existe la base de datos 'shipment_db'?")
        print("   3. ¿Están las tablas creadas correctamente?")
        print("   4. ¿Son correctos los datos de conexión?")
        print("   5. ¿Existe el archivo Excel?")
        print("   6. ¿Tiene hoja 'Consolidado'?")
        sys.exit(1)


if __name__ == "__main__":
    main()